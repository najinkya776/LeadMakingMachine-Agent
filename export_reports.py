#!/usr/bin/env python3
"""Export all PDF reports to Excel spreadsheet."""

import pdfplumber
import re
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def extract_report_data(pdf_path: Path) -> dict:
    """Extract data from a PDF report using pdfplumber."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

        if not text.strip():
            return None

        data = {
            'pdf_path': str(pdf_path),
            'pdf_filename': pdf_path.name,
        }

        # Extract business name from "Prepared for: <name>"
        name_match = re.search(r'Prepared for:\s*(.+?)(?:\n|Date)', text)
        if name_match:
            data['business_name'] = name_match.group(1).strip()

        # Extract date
        date_match = re.search(r'Date:\s*(.+?)(?:\n)', text)
        if date_match:
            data['report_date'] = date_match.group(1).strip()

        # Extract opportunity score and classification
        score_match = re.search(r'([A-Z]+)\s*\n\s*(\d+)\s*Opportunity Score', text)
        if score_match:
            data['classification'] = score_match.group(1).strip()
            data['opportunity_score'] = int(score_match.group(2))

        # Extract executive summary
        exec_match = re.search(r'Executive Summary\s*\n\s*(.+?)(?:\n\s*\n\s*[A-Z]|Recommendations)', text, re.DOTALL)
        if exec_match:
            data['executive_summary'] = exec_match.group(1).strip().replace('\n', ' ')[:500]

        # Extract phone from various formats
        phone_match = re.search(r'\+?[\d\s\-\(\)]{10,}', text)
        if phone_match:
            phone = phone_match.group().strip()
            if len(phone) >= 10:
                data['phone'] = phone

        # Extract email
        email_match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
        if email_match:
            data['email'] = email_match.group()

        # Extract address (location info)
        address_match = re.search(r'Ajmera Main Rd,?\s*[^,]+,\s*[^,]+,\s*[^,]+,\s*[^,]+,\s*[^,]+', text)
        if address_match:
            data['address'] = address_match.group().strip()

        # Extract website URL
        url_match = re.search(r'(https?://[^\s]+)', text)
        if url_match:
            data['website_url'] = url_match.group(1).strip()

        # Extract pricing from the table
        price_match = re.search(r'Basic[^\n]+\n([^\n]+)', text)
        if price_match:
            data['pricing_estimate'] = price_match.group(1).strip()

        # Extract category from location context (usually restaurant, shop etc)
        if 'restaurant' in text.lower() or 'veg' in text.lower():
            data['category'] = 'Restaurant'
        elif 'hotel' in text.lower():
            data['category'] = 'Hotel'
        elif 'clinic' in text.lower() or 'hospital' in text.lower():
            data['category'] = 'Healthcare'
        elif 'school' in text.lower() or 'college' in text.lower():
            data['category'] = 'Education'
        elif 'shop' in text.lower() or 'store' in text.lower():
            data['category'] = 'Retail'
        else:
            data['category'] = 'Business'

        # Extract lead type based on recommendations
        if 'redesign' in text.lower():
            data['lead_type'] = 'Redesign'
        elif 'mobile' in text.lower():
            data['lead_type'] = 'Mobile Optimization'
        elif 'seo' in text.lower() or 'optimization' in text.lower():
            data['lead_type'] = 'SEO'
        else:
            data['lead_type'] = 'New Website'

        # Pitch type
        if data['opportunity_score'] and data['opportunity_score'] >= 70:
            data['pitch_type'] = 'Premium'
        elif data['opportunity_score'] and data['opportunity_score'] >= 50:
            data['pitch_type'] = 'Standard'
        else:
            data['pitch_type'] = 'Basic'

        # Determine Google Maps/lead info
        data['source'] = 'Google Maps'

        # Extract recommendations count
        rec_matches = re.findall(r'\d+\.\s+(.+?)(?:\n|$)', text)
        data['recommendations_count'] = len(rec_matches)

        return data

    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")
        return None


def export_pdfs_to_excel(reports_dir: Path = None, output_path: Path = None):
    """Export all PDF reports to Excel."""
    reports_dir = Path("F:/My AI Project/website_pitcher/output/reports")

    if not reports_dir.exists():
        print(f"Reports directory not found: {reports_dir}")
        return None

    pdf_files = list(reports_dir.glob("report_*.pdf"))
    if not pdf_files:
        print("No PDF reports found")
        return None

    print(f"Found {len(pdf_files)} PDF reports")

    all_data = []
    for pdf_path in sorted(pdf_files):
        data = extract_report_data(pdf_path)
        if data:
            all_data.append(data)
            print(f"  {data.get('business_name', 'Unknown')[:40]}")

    if not all_data:
        print("No data could be extracted from PDFs")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Website Pitcher Reports"

    headers = [
        "Business Name", "Category", "Address", "Phone", "Email", "Website URL",
        "Opportunity Score", "Classification", "Lead Type", "Pitch Type",
        "Pricing Estimate", "Executive Summary", "Report Date", "PDF File"
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    high_score_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    med_score_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_score_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, data in enumerate(all_data, 2):
        def safe_val(key, default=""):
            return data.get(key, default)

        ws.cell(row=row_idx, column=1, value=safe_val('business_name')).border = thin_border
        ws.cell(row=row_idx, column=2, value=safe_val('category')).border = thin_border
        ws.cell(row=row_idx, column=3, value=safe_val('address')).border = thin_border
        ws.cell(row=row_idx, column=4, value=safe_val('phone')).border = thin_border
        ws.cell(row=row_idx, column=5, value=safe_val('email')).border = thin_border
        ws.cell(row=row_idx, column=6, value=safe_val('website_url')).border = thin_border

        opp_score = data.get('opportunity_score')
        score_cell = ws.cell(row=row_idx, column=7, value=opp_score)
        score_cell.border = thin_border
        score_cell.alignment = Alignment(horizontal='center')
        if opp_score:
            if opp_score >= 70:
                score_cell.fill = high_score_fill
            elif opp_score >= 50:
                score_cell.fill = med_score_fill
            else:
                score_cell.fill = low_score_fill

        ws.cell(row=row_idx, column=8, value=safe_val('classification')).border = thin_border
        ws.cell(row=row_idx, column=9, value=safe_val('lead_type')).border = thin_border
        ws.cell(row=row_idx, column=10, value=safe_val('pitch_type')).border = thin_border
        ws.cell(row=row_idx, column=11, value=safe_val('pricing_estimate')).border = thin_border

        summary_cell = ws.cell(row=row_idx, column=12, value=safe_val('executive_summary'))
        summary_cell.border = thin_border
        summary_cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws.cell(row=row_idx, column=13, value=safe_val('report_date')).border = thin_border
        ws.cell(row=row_idx, column=14, value=safe_val('pdf_filename')).border = thin_border

    column_widths = {
        'A': 35, 'B': 15, 'C': 45, 'D': 18, 'E': 30, 'F': 35,
        'G': 12, 'H': 12, 'I': 18, 'J': 12,
        'K': 22, 'L': 55, 'M': 15, 'N': 50,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:N{len(all_data) + 1}"

    output_dir = Path("F:/My AI Project/website_pitcher/output")
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"website_pitcher_all_reports_{timestamp}.xlsx"

    wb.save(output_path)
    print(f"\nExported {len(all_data)} reports to {output_path}")
    return output_path


if __name__ == "__main__":
    export_pdfs_to_excel()