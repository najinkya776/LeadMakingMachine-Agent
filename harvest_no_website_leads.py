#!/usr/bin/env python3
"""Harvest small businesses that have NO website and export their contact info to Excel.

This is the front-end of the lead pipeline. It scrapes Google Maps (via Apify),
keeps ONLY businesses that have no website / domain, and writes their contact
details into a single .xlsx file. That spreadsheet can later be fed back into the
main pipeline (qualify -> score -> pitch -> email) as a clean prospect list.

Target a whole region (expands to multiple cities for maximum coverage) or a
single custom location:

    python harvest_no_website_leads.py --region usa --count 120
    python harvest_no_website_leads.py --region uk --categories restaurant salon gym
    python harvest_no_website_leads.py --location "Pune, India" --count 50
    python harvest_no_website_leads.py --region australia --output output/au.xlsx --guess-emails

List available regions:
    python harvest_no_website_leads.py --list-regions

Notes:
    * Businesses without a website almost never expose a real email on Google Maps,
      so the `email` column is usually blank for this segment. Phone is the primary
      contact. Pass --guess-emails to fill blanks with a likely info@<name> guess
      (NOT verified - use with caution).
"""
import sys
import os
import argparse
from pathlib import Path
from datetime import datetime

# Fix Windows console encoding for Unicode characters
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).parent))

from dotenv import load_dotenv
load_dotenv()

from config.settings import ICP
from config.regions import REGIONS, get_locations, list_regions_grouped
from agents.scraper_primary import ScraperPrimaryAgent

# Columns written to the Excel sheet, in order. Pipeline-friendly: identity,
# then contact, then signals, then the region we sourced it from.
COLUMNS = [
    "business_name",
    "category",
    "phone",
    "email",
    "website",
    "address",
    "google_rating",
    "review_count",
    "reachability_score",
    "has_website",
    "region",
    "source_location",
    "source",
    "scraped_at",
]

# Website-presence filters.
FILTER_NO_WEBSITE = "no_website"    # keep only businesses with NO site (India angle)
FILTER_HAS_WEBSITE = "has_website"  # keep only businesses WITH a site (redesign/SEO angle)
FILTER_ALL = "all"                  # keep everything, tagged
WEBSITE_FILTERS = (FILTER_NO_WEBSITE, FILTER_HAS_WEBSITE, FILTER_ALL)

# Prefix used when auto-naming the output file.
_FILTER_PREFIX = {
    FILTER_NO_WEBSITE: "no_website",
    FILTER_HAS_WEBSITE: "has_website",
    FILTER_ALL: "all_leads",
}


# =============================================================================
# Core (reusable by the web UI)
# =============================================================================

def run_harvest(
    locations,
    categories=None,
    count=50,
    region_label="",
    guess_emails=False,
    output_path=None,
    website_filter=FILTER_NO_WEBSITE,
    progress=None,
):
    """Scrape the given locations, filter by website presence, write Excel.

    Args:
        locations: list of search-location strings (cities) to scrape.
        categories: business categories (defaults to all ICP industries).
        count: TOTAL target leads for the run, distributed across locations.
        region_label: human label stored on each row (e.g. "United States").
        guess_emails: fill blank emails with an unverified info@<name> guess.
        output_path: .xlsx destination (auto-named under output/ if None).
        website_filter: one of WEBSITE_FILTERS - keep no-website (default),
            has-website (redesign/SEO angle), or all (tagged).
        progress: optional callback(dict) for live status updates.

    Returns:
        dict summary: {output_path, total_scraped, kept, dropped, with_phone,
                       with_email, with_website, rows}
    """
    if website_filter not in WEBSITE_FILTERS:
        raise ValueError(f"website_filter must be one of {WEBSITE_FILTERS}")
    categories = categories or ICP["industries"]
    locations = list(locations)
    if not locations:
        raise ValueError("No locations to scrape.")

    def emit(**kw):
        if progress:
            progress(kw)

    # Distribute the total target across cities (floor of 5 each so small
    # regions still pull a meaningful sample).
    per_location = max(count // len(locations), 5)

    scraper = ScraperPrimaryAgent()
    all_pairs = []  # list of (lead, source_location)
    by_location = {}

    for idx, loc in enumerate(locations):
        emit(
            phase="scraping",
            message=f"Scraping {loc} ({idx + 1}/{len(locations)})",
            current=idx + 1,
            total=len(locations),
        )
        try:
            leads = scraper.run(location=loc, categories=categories, count=per_location)
        except Exception as e:  # never let one city kill the whole run
            emit(phase="warn", message=f"{loc}: {e}")
            leads = []
        all_pairs.extend((lead, loc) for lead in leads)
        by_location[loc] = len(leads)

    # Apply the website-presence filter.
    if website_filter == FILTER_HAS_WEBSITE:
        filtered = [(lead, loc) for lead, loc in all_pairs if lead.website_url]
    elif website_filter == FILTER_ALL:
        filtered = list(all_pairs)
    else:  # FILTER_NO_WEBSITE
        filtered = [(lead, loc) for lead, loc in all_pairs if not lead.website_url]

    # Dedupe on (name, phone) - the same business can appear across cities/categories.
    seen = set()
    unique = []
    for lead, loc in filtered:
        key = (lead.business_name.strip().lower(), (lead.phone or "").strip())
        if key in seen:
            continue
        seen.add(key)
        unique.append((lead, loc))

    emit(
        phase="exporting",
        message=f"{len(unique)} leads ({website_filter}) -> writing Excel",
    )

    rows = _to_rows(unique, region_label, guess_emails)

    if output_path is None:
        slug = region_label.lower().replace(" ", "_").replace("(", "").replace(")", "") or "leads"
        prefix = _FILTER_PREFIX.get(website_filter, "leads")
        output_path = os.path.join(
            "output", f"{prefix}_{slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )

    if rows:
        _write_excel(rows, output_path)

    summary = {
        "output_path": output_path if rows else None,
        "website_filter": website_filter,
        "total_scraped": len(all_pairs),
        "kept": len(rows),
        "dropped": len(all_pairs) - len(filtered),
        "with_phone": sum(1 for r in rows if r["phone"]),
        "with_email": sum(1 for r in rows if r["email"]),
        "with_website": sum(1 for r in rows if r["website"]),
        "by_location": by_location,
        "rows": rows,
    }
    emit(phase="done", message="Harvest complete", summary={k: v for k, v in summary.items() if k != "rows"})
    return summary


def _to_rows(pairs, region_label, guess_emails):
    """Convert (Lead, source_location) pairs into flat dict rows for the spreadsheet."""
    scraped_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    rows = []
    for lead, source_location in pairs:
        email = lead.email or ""
        if not email and guess_emails:
            slug = "".join(c for c in lead.business_name.lower() if c.isalnum())
            email = f"info@{slug}.com" if slug else ""

        rows.append({
            "business_name": lead.business_name,
            "category": lead.category or "",
            "phone": lead.phone or "",
            "email": email,
            "address": lead.address or "",
            "google_rating": lead.google_rating if lead.google_rating is not None else "",
            "review_count": lead.review_count if lead.review_count is not None else "",
            "reachability_score": lead.reachability_score if lead.reachability_score is not None else "",
            "website": lead.website_url or "",
            "has_website": "Yes" if lead.website_url else "No",
            "region": region_label,
            "source_location": source_location,
            "source": lead.source,
            "scraped_at": scraped_at,
        })
    return rows


def _write_excel(rows, output_path):
    """Write rows to a styled .xlsx file using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "No-Website Leads"

    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col])
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, col in enumerate(COLUMNS, start=1):
        max_len = len(col)
        for row in rows:
            val = str(row[col]) if row[col] is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    wb.save(output_path)


# =============================================================================
# CLI
# =============================================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Harvest no-website businesses and export contacts to Excel."
    )
    parser.add_argument(
        "--region",
        choices=list(REGIONS.keys()),
        default=None,
        help="Target region (expands to multiple cities). See --list-regions.",
    )
    parser.add_argument(
        "--location",
        default=None,
        help="Custom single location (overrides --region)",
    )
    parser.add_argument("--categories", nargs="+", default=None,
                        help="Business categories (default: all ICP industries)")
    parser.add_argument("--count", type=int, default=50,
                        help="Total target leads for the run (distributed across cities)")
    parser.add_argument("--website-filter", choices=list(WEBSITE_FILTERS),
                        default=FILTER_NO_WEBSITE,
                        help="Keep no_website (default), has_website (redesign/SEO angle), or all")
    parser.add_argument("--output", default=None, help="Output .xlsx path")
    parser.add_argument("--guess-emails", action="store_true",
                        help="Fill blank emails with an UNVERIFIED info@<business>.com guess")
    parser.add_argument("--list-regions", action="store_true",
                        help="Print available regions and exit")
    return parser.parse_args()


def _print_regions():
    grouped = list_regions_grouped()
    print("\nAvailable regions:\n")
    for tier in grouped.values():
        print(f"  {tier['label']}")
        for r in tier["regions"]:
            print(f"    {r['flag']} {r['key']:<14} {r['label']} ({r['city_count']} cities)")
        print()


def main():
    args = parse_args()

    if args.list_regions:
        _print_regions()
        return

    categories = args.categories or ICP["industries"]

    # Resolve locations: explicit --location wins, else --region, else default.
    if args.location:
        locations = [args.location]
        region_label = args.location
    elif args.region:
        locations = get_locations(args.region)
        region_label = REGIONS[args.region]["label"]
    else:
        # Default to Pune (preserves prior behaviour).
        locations = get_locations("pune")
        region_label = REGIONS["pune"]["label"]

    print("=" * 64)
    print("   No-Website Lead Harvester")
    print("=" * 64)
    print(f"   Region:     {region_label}")
    print(f"   Cities:     {len(locations)} -> {locations}")
    print(f"   Categories: {len(categories)}")
    print(f"   Count:      {args.count} (total, distributed across cities)")
    print(f"   Filter:     {args.website_filter}")
    print(f"   Guess emails: {args.guess_emails}")
    print()

    start = datetime.now()

    def progress(p):
        if p.get("message"):
            print(f"   [{p.get('phase','')}] {p['message']}")

    summary = run_harvest(
        locations=locations,
        categories=categories,
        count=args.count,
        region_label=region_label,
        guess_emails=args.guess_emails,
        output_path=args.output,
        website_filter=args.website_filter,
        progress=progress,
    )

    duration = (datetime.now() - start).total_seconds()
    print("\n" + "=" * 64)
    if summary["output_path"]:
        print(f"   Exported {summary['kept']} leads -> {summary['output_path']}")
        print(f"   With phone: {summary['with_phone']}/{summary['kept']}   "
              f"With email: {summary['with_email']}/{summary['kept']}")
    else:
        print("   No matching businesses found. Nothing exported.")
    drop_desc = {
        "no_website": "with websites",
        "has_website": "without websites",
        "all": "filtered out",
    }.get(args.website_filter, "filtered out")
    print(f"   Scraped {summary['total_scraped']} total, dropped {summary['dropped']} {drop_desc}.")
    print(f"   Done in {duration:.1f}s")
    print("=" * 64)


if __name__ == "__main__":
    main()
